"""Чтение присланных файлов — банковских выписок.

Разбирать формат конкретного банка кодом бессмысленно: у каждого
свои колонки, разделители и названия. Поэтому файл превращается в
текст, а разбор делает модель — она справляется с любой шапкой.
"""

import base64
import csv
import io
import logging
from dataclasses import dataclass
from pathlib import Path

logger = logging.getLogger("anfinances_bot.documents")

__all__ = [
    "MAX_CHARS",
    "MAX_IMAGE_BYTES",
    "MAX_PDF_BYTES",
    "Attachment",
    "DocumentUnreadableError",
    "read_image",
    "read_pdf",
    "read_spreadsheet",
    "read_statement",
]

# Сколько текста выписки отдаём модели. Больше — и ответ упрётся в
# окно контекста, а выписка на несколько тысяч строк всё равно
# просится разбитой по месяцам.
MAX_CHARS = 60_000

_ENCODINGS = ("utf-8-sig", "cp1251", "utf-16")


class DocumentUnreadableError(RuntimeError):
    """Файл не удалось прочитать как текст."""


def read_statement(path: Path) -> str:
    """Прочитать выписку в текст, что бы банк ни выгрузил.

    Российские банки любят cp1251 и точку с запятой, поэтому
    кодировки перебираются, а не угадываются по расширению.
    """
    raw = path.read_bytes()
    if not raw.strip():
        raise DocumentUnreadableError("Файл пустой.")

    text = None
    for encoding in _ENCODINGS:
        try:
            text = raw.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            continue
        else:
            break
    if text is None:
        raise DocumentUnreadableError(
            "Это не текстовый файл. Нужен CSV или TXT."
        )

    text = text.replace("\x00", "").strip()
    if not text:
        raise DocumentUnreadableError("Файл пустой.")

    if len(text) > MAX_CHARS:
        text = _head_rows(text)
    return text


def _head_rows(text: str) -> str:
    """Обрезать по строкам и честно сказать, что обрезано."""
    kept: list[str] = []
    total = 0
    for line in text.splitlines():
        if total + len(line) > MAX_CHARS:
            break
        kept.append(line)
        total += len(line) + 1
    logger.warning("Выписка обрезана до %s строк", len(kept))
    return "\n".join(kept) + (
        "\n\n[Файл обрезан: слишком длинный. Занеси показанное и "
        "попроси прислать остаток отдельным файлом.]"
    )


def sniff_delimiter(text: str) -> str:
    """Угадать разделитель — подсказка модели, а не разбор."""
    sample = "\n".join(text.splitlines()[:5])
    try:
        return str(csv.Sniffer().sniff(sample, delimiters=",;\t").delimiter)
    except (csv.Error, io.UnsupportedOperation):
        return ";"


# Скриншот чека, выписки или экрана банка. Клод принимает эти
# четыре типа; телеграм при пересылке фото отдаёт jpeg.
IMAGE_TYPES = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".webp": "image/webp",
    ".gif": "image/gif",
}

# Предел на изображение в запросе к модели.
MAX_IMAGE_BYTES = 5 * 1024 * 1024


def read_image(path: Path, filename: str = "") -> tuple[str, str]:
    """Прочитать картинку как пару «media_type, base64»."""
    raw = path.read_bytes()
    if not raw:
        raise DocumentUnreadableError("Файл пустой.")
    if len(raw) > MAX_IMAGE_BYTES:
        raise DocumentUnreadableError(
            "Картинка больше 5 МБ. Пришли скриншот поменьше."
        )

    media_type = _media_type(raw, filename)
    if media_type is None:
        raise DocumentUnreadableError(
            "Не похоже на картинку. Нужен jpg, png, webp или gif."
        )
    return media_type, base64.b64encode(raw).decode("ascii")


def _media_type(raw: bytes, filename: str) -> str | None:
    """Тип по сигнатуре файла, а не по расширению.

    Телеграм отдаёт фото без имени, а присланный «скриншот.pdf»
    именем врёт. Первые байты не врут.
    """
    if raw.startswith(b"\xff\xd8\xff"):
        return "image/jpeg"
    if raw.startswith(b"\x89PNG\r\n\x1a\n"):
        return "image/png"
    if raw.startswith(b"GIF8"):
        return "image/gif"
    if raw[:4] == b"RIFF" and raw[8:12] == b"WEBP":
        return "image/webp"
    return IMAGE_TYPES.get(Path(filename).suffix.lower())


MAX_PDF_BYTES = 30 * 1024 * 1024


def read_pdf(path: Path) -> str:
    """PDF в base64 — модель читает его сама.

    Вытаскивать текст из PDF кодом здесь плохая идея: банковские
    выписки свёрстаны таблицами, и извлечённый текст теряет
    привязку числа к строке. Модель видит страницу целиком.
    """
    raw = path.read_bytes()
    if not raw.startswith(b"%PDF"):
        raise DocumentUnreadableError("Файл повреждён: это не PDF.")
    if len(raw) > MAX_PDF_BYTES:
        raise DocumentUnreadableError(
            "PDF больше 30 МБ. Раздели выписку по месяцам."
        )
    return base64.b64encode(raw).decode("ascii")


def read_spreadsheet(path: Path) -> str:
    """XLSX в текст: строки таблицы через точку с запятой.

    Модель разберёт колонки сама, как и в CSV, — ей нужен только
    порядок значений в строке.
    """
    from openpyxl import load_workbook

    try:
        book = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise DocumentUnreadableError(
            "Не удалось открыть таблицу. Выгрузи из банка CSV."
        ) from exc

    lines: list[str] = []
    total = 0
    try:
        for sheet in book.worksheets:
            if len(book.worksheets) > 1:
                lines.append(f"# лист: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                if all(cell is None or cell == "" for cell in row):
                    continue
                line = ";".join(
                    "" if cell is None else str(cell) for cell in row
                )
                total += len(line) + 1
                if total > MAX_CHARS:
                    lines.append(
                        "\n[Таблица обрезана: слишком длинная. Занеси "
                        "показанное и попроси остаток отдельно.]"
                    )
                    return "\n".join(lines)
                lines.append(line)
    finally:
        book.close()

    if not lines:
        raise DocumentUnreadableError("В таблице нет строк.")
    return "\n".join(lines)


@dataclass(frozen=True)
class Attachment:
    """Прочитанное вложение в виде, пригодном для модели.

    Ровно одно из полей заполнено: картинка парой «тип, base64»,
    PDF — base64 страниц, таблица или CSV — текстом.
    """

    name: str
    image: tuple[str, str] | None = None
    pdf: str | None = None
    text: str | None = None

    @property
    def kind(self) -> str:
        if self.image is not None:
            return "скриншот"
        if self.pdf is not None:
            return "PDF"
        return "таблица"
